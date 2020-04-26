import numpy as np
import matplotlib.pyplot as plt
import joblib, dill, pickle
import openpyxl
from deap import base, creator, tools
from collections import Counter
from dill import Pickler
from joblib.pool import CustomizablePicklingQueue
from io import BytesIO
from pickle import HIGHEST_PROTOCOL
from joblib import Parallel, delayed

from own_packages.base_opt import base_opt
from own_packages.basic_runs import execute_base_opt
from own_packages.others import create_results_directory, print_array_to_excel
from own_packages.ga import create_hparams, own_ea, Own_HallOfFame


def ga_opt(loss, mode, efficiency, consumer_class, hparams, results_dir, save_path,
           cost_factor, init_fraction):
    total_customers = len(consumer_class)
    if loss == 'par':
        creator.create("FitnessMax", base.Fitness, weights=(-1.0, -1e6,))
    elif loss == 'normal':
        creator.create("FitnessMax", base.Fitness, weights=(-1e8, -1.0,))
    elif loss == 'test':
        creator.create("FitnessMax", base.Fitness, weights=(-1, -1e8,))
    else:
        raise KeyError('Loss mode not one of the accepted types.')
    creator.create("Individual", list, fitness=creator.FitnessMax)
    global kk, lm
    kk = 1

    # Initial initialisation of lm matrix
    lm = []
    for consumer in consumer_class:
        lm.append(consumer.xn_init)

    lm = np.squeeze(np.array(lm)).T

    def atest_cost_function(individual):
        counter = Counter(individual)
        return(counter[0], counter[1],)

    def eval_total_cost(individual):
        esd_cost = 2.6* cost_factor
        sp_cost = 1.9726 * cost_factor
        total_cost, _, par = base_opt(lm=lm, consumer_class=consumer_class, esd_assignment=individual,
                                      full_iter=hparams['full_iter'], efficiency=efficiency,
                                      plot_mode=False,
                                      plot_dir=results_dir + '/plots', plot_subname=kk)
        # 2.6 is daily depreciation cost
        counter = Counter(individual)
        total_cost = counter[1] * esd_cost + counter[2] * (esd_cost+sp_cost) + total_cost
        return (total_cost,(par-1),)

    def eval_total_cost_with_par(individual):
        esd_cost = 2.6* cost_factor
        sp_cost = 1.9726 * cost_factor
        total_cost, _, par = base_opt(lm=lm, consumer_class=consumer_class, esd_assignment=individual,
                                      full_iter=hparams['full_iter'],efficiency=efficiency,
                                      plot_mode=False,
                                      plot_dir=results_dir + '/plots', plot_subname=kk)
        counter = Counter(individual)
        total_cost = counter[1] * esd_cost + counter[2] * (esd_cost+sp_cost) + total_cost
        return (total_cost, (par - 1),)

    toolbox = base.Toolbox() #
    if mode == 'esd':
        # toolbox.register("attr_bool", random.randint, 0, 1)
        toolbox.register("attr_bool", np.random.choice,np.arange(0, 2), p=init_fraction)
        toolbox.register("individual", tools.initRepeat, creator.Individual, toolbox.attr_bool, n=total_customers)
        toolbox.register("population", tools.initRepeat, list, toolbox.individual)
        if loss == 'normal':
            toolbox.register("evaluate", eval_total_cost)
        elif loss == 'par':
            toolbox.register("evaluate", eval_total_cost_with_par)
        elif loss == 'test':
            toolbox.register("evaluate", atest_cost_function)
        else:
            raise KeyError('Type selected is not one of the valid cost function type for the GA opt.')
        toolbox.register("mate", tools.cxTwoPoint)
        toolbox.register("mutate", tools.mutFlipBit, indpb=0.05)
        toolbox.register("select", tools.selTournament, tournsize=3)
        toolbox.register("map", mymap)
    elif mode == 'sp':
        #toolbox.register("attr_int", random.randint, 0, 2)
        toolbox.register("attr_int", np.random.choice,np.arange(0, 3), p=init_fraction)
        toolbox.register("individual", tools.initRepeat, creator.Individual, toolbox.attr_int, n=total_customers)
        toolbox.register("population", tools.initRepeat, list, toolbox.individual)
        if loss == 'normal':
            toolbox.register("evaluate", eval_total_cost)
        elif loss == 'par':
            toolbox.register("evaluate", eval_total_cost_with_par)
        elif loss == 'test':
            toolbox.register("evaluate", atest_cost_function)
        else:
            raise KeyError('Type selected is not one of the valid cost function type for the GA opt.')
        toolbox.register("mate", tools.cxTwoPoint)
        toolbox.register("mutate", tools.mutUniformInt, low=0, up=2, indpb=0.05)
        toolbox.register("select", tools.selTournament, tournsize=3)
        toolbox.register("map", mymap)
    else:
        raise KeyError('Mode selected is not one of the valid GA mode.')

    # Logging
    stats = tools.Statistics(key=lambda ind: ind.fitness.values)
    stats.register("avg", np.mean, axis=0)
    stats.register("std", np.std, axis=0)
    stats.register("min", np.min, axis=0)
    stats.register("max", np.max, axis=0)
    hof = Own_HallOfFame(1)

    pop = toolbox.population(n=hparams['n_pop'])
    pop, logbook = own_ea(pop, toolbox,
                          stats=stats, halloffame=hof,
                          tourn_size=3,cxpb=0.5, mutpb=0.2, ngen=hparams['n_gen'], verbose=True)

    print(hof[-1])

    # Plotting
    gen = logbook.select("gen")
    fit_mins = logbook.select("min")
    cost_min = [min[0] for min in fit_mins]
    par_min = [min[1] for min in fit_mins]
    fit_avg = logbook.select("avg")
    cost_avg = [avg[0] for avg in fit_avg]
    par_avg = [avg[1] for avg in fit_avg]
    fit_max = logbook.select("max")
    cost_max = [max[0] for max in fit_max]
    par_max = [max[1] for max in fit_max]

    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, cost_min, label="Min Cost")
    line2 = ax1.plot(gen, cost_avg, label="Avg Cost")
    line3 = ax1.plot(gen, cost_max, label="Max Cost")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Total Generation Cost")
    plt.savefig('{}/plots/GA_opt_cost.png'.format(results_dir), bbox_inches="tight")

    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, cost_min, label="Min Cost")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Total Generation Cost")
    plt.savefig('{}/plots/GA_opt_min_only_cost.png'.format(results_dir), bbox_inches="tight")

    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, par_min, label="Min PAR")
    line2 = ax1.plot(gen, par_avg, label="Avg PAR")
    line3 = ax1.plot(gen, par_max, label="Max PAR")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("PAR")
    plt.savefig('{}/plots/GA_opt_par.png'.format(results_dir), bbox_inches="tight")

    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, par_min, label="Min PAR")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("PAR")
    plt.savefig('{}/plots/GA_opt_min_only_PAR.png'.format(results_dir), bbox_inches="tight")

    fig, ax1 = plt.subplots()
    ax2 = ax1.twinx()
    ax1.plot(gen, cost_min, 'g-')
    ax2.plot(gen, par_min, 'b-')
    ax1.set_xlabel('Generation')
    ax1.set_ylabel('Total Cost', color='g')
    ax2.set_ylabel('PAR', color='b')
    plt.savefig('{}/plots/GA_opt_min_only_both.png'.format(results_dir), bbox_inches="tight")

    fig, ax1 = plt.subplots()
    ax2 = ax1.twinx()
    ax1.plot(gen, cost_min, 'g-')
    ax2.plot(gen, par_min, 'g--')
    ax1.plot(gen, cost_avg, 'r-')
    ax2.plot(gen, par_avg, 'r--')
    ax1.plot(gen, cost_max, 'b-')
    ax2.plot(gen, par_max, 'b--')
    ax1.set_xlabel('Generation')
    ax1.set_ylabel('Total Cost (Solid)')
    ax2.set_ylabel('PAR (Dashed)')
    plt.savefig('{}/plots/GA_opt_both.png'.format(results_dir), bbox_inches="tight")

    best_result = hof[-1]
    print('Total ESD Allocated: {}\nList of best allocation for each GA generation:\n{}'.format(
        sum(best_result), fit_mins))

    # Printing to excel
    execute_base_opt(esd_av=best_result,
                     save_path=save_path, dir_name=results_dir, excel_name='optimal_base_opt',
                     cf=cost_factor, efficiency=efficiency,
                     full_iter=2,
                     save_mode=False)
    excel_name = results_dir + '/results.xlsx'
    wb = openpyxl.Workbook()
    wb.save(excel_name)
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]

    # Writing other subset split, instance per run, and bounds
    print_array_to_excel(['loss', 'mode', 'n_gen', 'n_pop', 'n_iter'], (1, 1), ws, axis=1)
    print_array_to_excel([loss, mode, hparams['n_gen'], hparams['n_pop'], hparams['full_iter']], (2, 1), ws, axis=1)
    row = 2
    ws.cell(row+1,1).value = 'Best ESD Allocation'
    print_array_to_excel(best_result, (row+2, 1), ws, axis=1)
    ws.cell(row+3, 1).value = 'Total ESD Allocated'
    esd_allocated = sum(x==1 for x in best_result)
    ws.cell(row+3, 2).value = esd_allocated
    ws.cell(row+4, 1).value = 'Alpha'
    ws.cell(row+4, 2).value = esd_allocated/total_customers
    ws.cell(row + 3, 3).value = 'Total ESD Allocated'
    sp_allocated = sum(x == 2 for x in best_result)
    ws.cell(row+3, 4).value = sp_allocated
    ws.cell(row+4, 3).value = 'Beta'
    ws.cell(row+4, 4).value = sp_allocated/total_customers
    row += 1
    ws.cell(row+4, 1).value = 'Total Cost Statistics'
    ws.cell(row+5, 1).value = 'Gen'
    ws.cell(row+6, 1).value = 'Min'
    ws.cell(row+7, 1).value = 'Avg'
    ws.cell(row+8, 1).value = 'Max'
    print_array_to_excel(list(range(1,len(fit_mins)+1)), (row+5, 2), ws, axis=1)
    print_array_to_excel(cost_min, (row+6, 2), ws, axis=1)
    print_array_to_excel(cost_avg, (row+7, 2), ws, axis=1)
    print_array_to_excel(cost_max, (row+8, 2), ws, axis=1)
    row+=9
    ws.cell(row+4, 1).value = 'PAR Statistics'
    ws.cell(row+5, 1).value = 'Gen'
    ws.cell(row+6, 1).value = 'Min'
    ws.cell(row+7, 1).value = 'Avg'
    ws.cell(row+8, 1).value = 'Max'
    print_array_to_excel(list(range(1,len(fit_mins)+1)), (row+5, 2), ws, axis=1)
    print_array_to_excel(par_min, (row+6, 2), ws, axis=1)
    print_array_to_excel(par_avg, (row+7, 2), ws, axis=1)
    print_array_to_excel(par_max, (row+8, 2), ws, axis=1)

    wb.create_sheet('av')
    ws = wb['av']
    ws.cell(1,1).value = 'av'
    print_array_to_excel(best_result, (2,1), ws=ws, axis=0)

    wb.save(excel_name)
    wb.close()


if __name__ == '__main__':

    ####################################################################################################################
    # First half is the prep for multi-processing GA opt
    # Ignore this part, it is just to make the multi-processing with DEAP work.
    joblib.parallel.pickle = dill
    joblib.pool.dumps = dill.dumps
    joblib.pool.Pickler = Pickler

    class CustomizablePickler(Pickler):
        """Pickler that accepts custom reducers.
        HIGHEST_PROTOCOL is selected by default as this pickler is used
        to pickle ephemeral datastructures for interprocess communication
        hence no backward compatibility is required.
        `reducers` is expected expected to be a dictionary with key/values
        being `(type, callable)` pairs where `callable` is a function that
        give an instance of `type` will return a tuple `(constructor,
        tuple_of_objects)` to rebuild an instance out of the pickled
        `tuple_of_objects` as would return a `__reduce__` method. See the
        standard library documentation on pickling for more details.
        """

        # We override the pure Python pickler as its the only way to be able to
        # customize the dispatch table without side effects in Python 2.6
        # to 3.2. For Python 3.3+ leverage the new dispatch_table
        # feature from http://bugs.python.org/issue14166 that makes it possible
        # to use the C implementation of the Pickler which is faster.

        def __init__(self, writer, reducers=None, protocol=HIGHEST_PROTOCOL):
            Pickler.__init__(self, writer, protocol=protocol)
            if reducers is None:
                reducers = {}
            # Make the dispatch registry an instance level attribute instead of
            # a reference to the class dictionary under Python 2
            self.dispatch = Pickler.dispatch.copy()
            for type, reduce_func in reducers.items():
                self.register(type, reduce_func)

        def register(self, type, reduce_func):
            if hasattr(Pickler, 'dispatch'):
                # Python 2 pickler dispatching is not explicitly customizable.
                # Let us use a closure to workaround this limitation.
                def dispatcher(self, obj):
                    reduced = reduce_func(obj)
                    self.save_reduce(obj=obj, *reduced)

                self.dispatch[type] = dispatcher
            else:
                self.dispatch_table[type] = reduce_func

    joblib.pool.CustomizablePickler = CustomizablePickler

    def _make_methods(self):
        self._recv = recv = self._reader.recv
        racquire, rrelease = self._rlock.acquire, self._rlock.release

        def get():
            racquire()
            try:
                return recv()
            finally:
                rrelease()

        self.get = get

        def send(obj):
            buffer = BytesIO()
            CustomizablePickler(buffer, self._reducers).dump(obj)
            self._writer.send_bytes(buffer.getvalue())

        self._send = send

        if self._wlock is None:
            # writes to a message oriented win32 pipe are atomic
            self.put = send
        else:
            wlock_acquire, wlock_release = (
                self._wlock.acquire, self._wlock.release)

            def put(obj):
                wlock_acquire()
                try:
                    return send(obj)
                finally:
                    wlock_release()

            self.put = put

    CustomizablePicklingQueue._make_methods = _make_methods

    def mymap(f, *iters):
        creator.create("FitnessMax", base.Fitness, weights=(-1.0,))
        creator.create("Individual", list, fitness=creator.FitnessMax)
        return Parallel(n_jobs=-1)(delayed(f)(*args) for args in zip(*iters))

    ####################################################################################################################
    # Second half is the actual GA opt
    def ga_run(results_dir, save_path, loss, mode, gen, pop, iter,
               efficiency, cost_factor, init_fraction):
        '''
        To run the GA optimization to find the optimal allocation of SP and ES in the population.
        Run the run_preprocess.py script first to convert the excel data into a consumer_class.dat file which will be
        read from the save_path argument.

        The optimizer works in the following phases:
        1) The loss function is defined: normal or esd mode.
        2) The population of individuals are created. Each individual is a vector whose elements are 0 (no units),
        1 (batteries only), or 2 (batteries + solar panels)
        3) The GA optimizer then iterates for the specified number of generations.
        4) In each iteration, each individual in the population undergoes a base opt, which is simulating the
        operational phase where the DSM is operating to reduce bills and PAR.
        5) After the base opt, the score which is the avg. daily bills is stored as well as the PAR.
        6) The population is ran through a tournament of size 3, where the top 1 out of 3 individual survives and
        this tournamnet iterates through the population.
        7) Then there is the crossover and mutation which then concludes one generation.
        8) The rest of the code is for plotting the convergence plot and printing the results into excel.

        :param results_dir: Name of the results folder to create.
        :param save_path: Name of the consumer_class dat file that was saved from main.py, case == 1
        :param loss: Type of objective function.
        'normal' = only consider bill savings
        'par' = objective to set PAR=1, then consider bill savings.
        :param mode: Whether to have ESD only or ESD+SP
        'esd' = esd only
        'sp' = esd and sp together
        :param gen: Number of GA generations
        :param pop: Number of different combinations in one generation
        :param iter: Number of iteration for the base opt. Usually set to 2.
        :param efficiency: Efficiency factor for storage device. 1 means 100%
        :param cost_factor: 0.9 means cost of SP and batteries are 90% of original price
        :param init_fraction: Fraction of no units / ES / SP for initial guess of first generation
        E.g.: 2/6, 0, 4/6 ==> 2/6 will have no units, 0 will have ES, 4/6 will have SP on average for the first pop.
        :return:
        '''
        dir = create_results_directory(results_dir)
        with open(save_path, "rb") as f:
            consumer_class = pickle.load(f)
        hparams = create_hparams(n_pop=pop, n_gen=gen, full_iter=iter)
        ga_opt(loss=loss, mode=mode, efficiency=efficiency, consumer_class=consumer_class, hparams=hparams,
               results_dir=dir, save_path=save_path,
               cost_factor=cost_factor, init_fraction=init_fraction)

    ga_run(results_dir='./results/ga_opt_results_60c_s100',
           save_path='./save/consumer_class/60c_s100.dat', loss='normal', mode='sp', gen=5, pop=10, iter=2,
           efficiency=None, cost_factor=1,
           init_fraction=[2/6, 0, 4/6])



