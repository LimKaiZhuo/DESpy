import numpy as np
import pandas as pd
import openpyxl
import pickle
from collections import Counter

from own_packages.preprocess import read_consumer_excel
from own_packages.base_opt import base_opt
from own_packages.others import create_results_directory, print_array_to_excel, print_df_to_excel


def initialize_consumer_class(excel_path, save_path):
    return read_consumer_excel(excel_path, save_path=save_path)


def save_consumer_class_to_file(excel_path, solar_path, name, sigma, save_path, k):
    '''
    Reads the excel file which contains the consumer data.

    There are 6 worksheets for each consumer class (high, middle, low).

    S_am: A 2D array of dimensions i rows × j columns, where
    i = number of number of shiftable appliances and
    j = number of households in each consumer group m (1 to 3).
    Each element denotes the earliest starting 1-hour period (a) of shiftable appliance i in household j
    (where 0 = 12mn - 1am, 1 = 1am - 2am etc.).

    S_bm: A 2D array of dimensions i rows × j columns, where
    i = number of number of shiftable appliances and
    j = number of households in each consumer group m (1 to 3).
    Each element denotes the latest ending 1-hour period (b) of shiftable appliance i in household j
    (where 0 = 12mn - 1am, 1 = 1am - 2am etc.).

    S_cm: A 2D array of dimensions i rows × j columns, where
    i = number of number of shiftable appliances and
    j = number of households in each consumer group m (1 to 3).
    Each element denotes the daily usage duration (C) of shiftable appliance i in household j
    (in hours).

    S_dm: A 2D array of dimensions i rows × j columns, where
    i = number of number of shiftable appliances and
    j = number of households in each consumer group m (1 to 3).
    Each element denotes the daily total load (D) of shiftable appliance i in household j
    (in watt-hours or Wh).

    NSm: A 2D array of dimensions j rows × h columns, where
    j = number of households in each consumer group m (= 30) and
    h = number of 1-hour periods in a day (= 24).
    Each element denotes the hourly load profile of all nonshiftable appliances in household j combined
    (in watt-hours or Wh).

    S2NSm: A 2D array of dimensions j rows × h columns, where
    j = number of households in each consumer group m (= 30) and
    h = number of 1-hour periods in a day (= 24).
    Each element denotes the hourly load profile of all "previously shiftable" appliances in
    household j combined (in watt-hours or Wh), which is used to generate an alternative nonshiftable
    array if household j was to opt out of demand scheduling program (such that all shiftable loads
    are converted to nonshiftable). In this case, without loss of generality the usage of each
    "previously shiftable" appliance i begins at 1-hour period 'a' for duration 'C'.
    '''
    read_consumer_excel(consumer_excel_file=excel_path, solar_excel_file=solar_path,
                        sigma=sigma, k=k,
                        save_path=save_path)
    with open(save_path, "rb") as f:
        consumer_class = pickle.load(f)

    execute_base_opt(esd_av=np.zeros(len(consumer_class)), cf=1, save_path=save_path,
                     new_save_path=save_path,
                     save_mode=True,
                     dir_name=None, excel_name='base_opt_{}'.format(name), full_iter=3)


def execute_base_opt(esd_av, cf, save_path,
                     dir_name,
                     full_iter=2,excel_name='base_opt_results',
                     save_mode=False, efficiency=None, new_save_path=None):
    '''
    Runs the base opt which is the operational phase and prints the results to an excel file.
    '''
    with open(save_path, "rb") as f:
        consumer_class = pickle.load(f)

    # Initial initialisation of lm matrix
    lm = []
    for consumer in consumer_class:
        lm.append(consumer.xn_init)
        consumer.esd_av = esd_av
    lm = np.squeeze(np.array(lm)).T
    if dir_name:
        tc_store, _, par_store = base_opt(lm=lm, consumer_class=consumer_class, esd_assignment=esd_av,
                                          efficiency=efficiency,
                                          plot_mode=True, plot_dir=dir_name + '/plots',
                                          full_iter=full_iter, full_return=True)
    else:
        tc_store, _, par_store = base_opt(lm=lm, consumer_class=consumer_class, esd_assignment=esd_av,
                                          efficiency=efficiency,
                                          plot_mode=False, plot_dir=None,
                                          full_iter=full_iter, full_return=True)

    # Printing to excel
    if dir_name:
        excel_name = '{}/{}.xlsx'.format(dir_name, excel_name)
        wb = openpyxl.Workbook()

        # mainsheet formatting
        mainsheet = wb.sheetnames[-1]
        ws = wb[mainsheet]

        ws.cell(1, 1).value = 'No. Customer'
        ws.cell(1, 3).value = 'k'
        ws.cell(1, 4).value = 'cf'
        ws.cell(1, 5).value = 'sigma'
        ws.cell(2, 1).value = len(consumer_class)
        ws.cell(2, 3).value = consumer_class[0].k
        ws.cell(2, 4).value = cf
        ws.cell(2, 5).value = consumer_class[0].sigma
        print_array_to_excel(np.arange(len(consumer_class)) + 1, (3, 3), ws, axis=1)
        print_array_to_excel(['Consumer'], (3, 1), ws, axis=0)
        print_array_to_excel(['Bills'], (4, 1), ws, axis=0)
        print_array_to_excel(['Iteration'], (4, 2), ws, axis=0)
        print_array_to_excel(np.arange(full_iter + 1) + 1, (5, 2), ws, axis=0)
        row1 = 4 + full_iter + 1 + 1 + 1
        print_array_to_excel(['Final qss'], (row1 + 1, 1), ws, axis=0)
        print_array_to_excel(['-'], (row1 + 1, 2), ws, axis=0)
        print_array_to_excel(['Final xnb'], (row1 + 2, 1), ws, axis=0)
        print_array_to_excel(np.arange(24), (row1 + 2, 2), ws, axis=0)
        row2 = row1 + 2 + 24 + 1
        print_array_to_excel(['Final xn'], (row2 + 1, 1), ws, axis=0)
        print_array_to_excel(np.arange(24), (row2 + 1, 2), ws, axis=0)

        class_counter = Counter(consumer_class[-1].esd_av)
        wb.create_sheet('Summary')
        ws = wb[wb.sheetnames[-1]]
        ws.cell(1, 1).value = 'Iteration'
        print_array_to_excel(list(range(1, len(tc_store) + 1)), (1, 2), ws=ws, axis=1)
        ws.cell(2, 1).value = 'Total Cost w/o units'
        print_array_to_excel(tc_store, (2, 2), ws=ws, axis=1)
        tc_store_with_units = [tc + (class_counter[1] + class_counter[2]) * 2.6 * cf + class_counter[2] * 1.9726 * cf
                               for tc in tc_store]
        ws.cell(3, 1).value = 'Total Cost with units'
        print_array_to_excel(tc_store_with_units, (3, 2), ws=ws, axis=1)
        ws.cell(4, 1).value = 'PAR'
        print_array_to_excel(par_store, (4, 2), ws=ws, axis=1)
        df = pd.DataFrame(data=np.array([[consumer_class[-1].tcost_init, 0,
                                          consumer_class[-1].tcost_init / len(consumer_class), consumer_class[-1].par_init],
                                         [tc_store[-1], 100 * (1 - tc_store[-1] / consumer_class[-1].tcost_init),
                                          tc_store[-1] / len(consumer_class), par_store[-1]],
                                         [tc_store_with_units[-1],
                                          100 * (1 - tc_store_with_units[-1] / consumer_class[-1].tcost_init),
                                          tc_store_with_units[-1] / len(consumer_class), par_store[-1]]]),
                          columns=['Total Cost', 'Savings %', 'Average Bills', 'PAR'],
                          index=['Initial', 'Final w/o Units', 'Final with Units'])
        print_df_to_excel(df=df, ws=ws, start_row=5, start_col=1)
        skip_row = 6
        ws.cell(4 + skip_row, 1).value = 'ESD Allocation Vector'
        print_array_to_excel(array=list(range(1, len(consumer_class) + 1)), first_cell=(5 + skip_row, 1), ws=ws, axis=1)
        print_array_to_excel(array=consumer_class[-1].esd_av, first_cell=(6 + skip_row, 1), ws=ws, axis=1)
        ws.cell(7 + skip_row, 1).value = 'Total ESD'
        ws.cell(8 + skip_row, 1).value = 'Total SP'
        ws.cell(7 + skip_row, 2).value = class_counter[1]
        ws.cell(8 + skip_row, 2).value = class_counter[2]
        type1_counter = Counter(consumer_class[-1].esd_av[:20])
        type2_counter = Counter(consumer_class[-1].esd_av[20:40])
        type3_counter = Counter(consumer_class[-1].esd_av[40:])
        ws.cell(7 + skip_row, 3).value = 'Class 1 ESD'
        ws.cell(8 + skip_row, 3).value = 'Class 1 SP'
        ws.cell(7 + skip_row, 4).value = type1_counter[1]
        ws.cell(8 + skip_row, 4).value = type1_counter[2]
        ws.cell(7 + skip_row, 5).value = 'Class 2 ESD'
        ws.cell(8 + skip_row, 5).value = 'Class 2 SP'
        ws.cell(7 + skip_row, 6).value = type2_counter[1]
        ws.cell(8 + skip_row, 6).value = type2_counter[2]
        ws.cell(7 + skip_row, 7).value = 'Class 3 ESD'
        ws.cell(8 + skip_row, 7).value = 'Class 3 SP'
        ws.cell(7 + skip_row, 8).value = type3_counter[1]
        ws.cell(8 + skip_row, 8).value = type3_counter[2]

        bn_cs_store = []
        # Looping through customers

        for idx, consumer in enumerate(consumer_class):
            col = idx + 3
            ws = wb[mainsheet]
            print_array_to_excel(consumer.bn_store, (5, col), ws, axis=0)
            bn_cs = 100 * (1 - consumer.bn_store[-1] / consumer.bn_store[0])
            print_array_to_excel([bn_cs], (row1 - 1, col), ws, axis=0)
            bn_cs_store.append(bn_cs)

            try:
                print_array_to_excel([consumer.qss_store[-1]], (row1 + 1, col), ws, axis=0)
            except:
                pass
            try:
                print_array_to_excel(consumer.xnb_store[-1], (row1 + 2, col), ws, axis=2)
            except:
                pass
            print_array_to_excel(consumer.xn_store[-1], (row2 + 1, col), ws, axis=2)

            new_c_sheet = 'c{}'.format(idx + 1)
            wb.create_sheet(new_c_sheet)
            ws = wb[new_c_sheet]

            print_array_to_excel(['xn'], (1, 1), ws, axis=0)
            print_array_to_excel(np.arange(full_iter + 1) + 1, (1, 2), ws, axis=1)
            print_array_to_excel(np.arange(24), (2, 1), ws, axis=0)
            print_array_to_excel(np.squeeze(np.array(consumer.xn_store), axis=2).T, (2, 2), ws, axis=2)

            row3 = 27
            print_array_to_excel(['xnb'], (row3, 1), ws, axis=0)
            print_array_to_excel(np.arange(full_iter) + 1, (row3, 2), ws, axis=1)
            print_array_to_excel(np.arange(24), (row3 + 1, 1), ws, axis=0)
            try:
                print_array_to_excel(np.squeeze(np.array(consumer.xnb_store), axis=2).T, (row3 + 1, 2), ws, axis=2)
            except:
                pass

            row4 = 53
            print_array_to_excel(['Final t'], (row4, 1), ws, axis=0)
            print_array_to_excel(np.arange(consumer.app_count) + 1, (row4, 2), ws, axis=1)
            print_array_to_excel(np.arange(24), (row4 + 1, 1), ws, axis=0)
            print_array_to_excel(consumer.t_store[-1], (row4 + 1, 2), ws, axis=2)


        ws = wb[mainsheet]
        print_array_to_excel(['bn cost savings (%)'], (row1 - 1, 2), ws, axis=0)
        print_array_to_excel(['bn avg cost savings'], (row1, 2), ws, axis=0)
        print_array_to_excel([np.average(np.array(bn_cs_store))], (row1, 3), ws, axis=0)

        wb.save(excel_name)
        wb.close()

    if save_mode:
        with open(new_save_path, "wb") as f:
            pickle.dump(consumer_class, f)

    return consumer_class
